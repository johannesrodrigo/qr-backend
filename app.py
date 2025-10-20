import hmac, hashlib, base64, io, os, time
from typing import Optional
from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import JSONResponse
import requests
from openpyxl import load_workbook

app = FastAPI(title="QR Backend (sin pandas)", version="1.0.0")

# ===== utilidades =====

def od_to_download(url: str) -> str:
    """Convierte el link de 'Compartir' de OneDrive/SharePoint a URL de descarga directa."""
    if "sharepoint.com" in url or "my.sharepoint.com" in url:
        # SharePoint: agregar &download=1
        return url + ("&download=1" if "?" in url else "?download=1")
    if "1drv.ms" in url or "onedrive.live.com" in url:
        # Live/short: agregar &download=1 (simple)
        return url + ("&download=1" if "?" in url else "?download=1")
    return url

def secure_eq(a: str, b: str) -> bool:
    try:
        return hmac.compare_digest(a.encode(), b.encode())
    except Exception:
        return False

def sign(doc: str, secret_key: str) -> str:
    mac = hmac.new(secret_key.encode(), doc.encode(), hashlib.sha256).digest()
    return base64.urlsafe_b64encode(mac).decode().rstrip("=")

def verify(doc: str, t: str, secret_key: str) -> bool:
    try:
        expected = sign(doc, secret_key)
        # normalizar padding de base64 urlsafe
        return secure_eq(t, expected)
    except Exception:
        return False

def normalize(s: Optional[str]) -> str:
    if s is None:
        return ""
    return " ".join(str(s).strip().split()).upper()

# ===== lectura Excel SIN pandas =====

def read_driver_from_excel(
    xls_bytes: bytes,
    sheet_name: Optional[str],
    header_row_1based: int,
    dni_value: str,
) -> dict:
    """
    Lee el Excel y devuelve el registro del conductor por DNI de la COLUMNA E,
    y extrae: D (NOMBRES Y APELLIDOS), E (DNI / CE), AF (FECHA DE VIGENCIA ...),
    AG (ESTATUS DE PROCESO DE HABILITACION) usando los encabezados.
    """
    wb = load_workbook(io.BytesIO(xls_bytes), data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    H = header_row_1based
    headers = {}
    # construir mapa de encabezados (posición -> texto normalizado)
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=H, column=col).value
        headers[col] = normalize(val)

    # buscamos las columnas por NOMBRE de encabezado (tolerante a espacios)
    need = {
        "NOMBRES Y APELLIDOS": None,        # col D
        "DNI / CE": None,                   # col E
        "FECHA DE VIGENCIA DE HABILITACIÓN DE LICENCIA INTERNA": None,  # col AF
        "ESTATUS DE PROCESO DE HABILITACION": None,                      # col AG
    }

    for col, text in headers.items():
        for k in list(need.keys()):
            # match exacto tras normalizar
            if text == normalize(k):
                need[k] = col

    missing = [k for k, v in need.items() if v is None]
    if missing:
        raise HTTPException(
            status_code=400,
            detail={"error": "Faltan columnas requeridas por encabezado", "missing": missing, "encabezados": list(headers.values())}
        )

    col_name = need["NOMBRES Y APELLIDOS"]
    col_dni  = need["DNI / CE"]
    col_fvig = need["FECHA DE VIGENCIA DE HABILITACIÓN DE LICENCIA INTERNA"]
    col_stat = need["ESTATUS DE PROCESO DE HABILITACION"]

    dni_value_norm = normalize(dni_value)

    # recorrer filas de datos
    for row in range(H + 1, ws.max_row + 1):
        cell_dni = normalize(ws.cell(row=row, column=col_dni).value)
        if cell_dni == dni_value_norm:
            nombres = ws.cell(row=row, column=col_name).value
            fecha_vig = ws.cell(row=row, column=col_fvig).value
            estatus = ws.cell(row=row, column=col_stat).value
            return {
                "NOMBRES_Y_APELLIDOS": str(nombres or "").strip(),
                "DNI_CE": dni_value_norm,
                "FECHA_VIGENCIA_LICENCIA_INTERNA": str(fecha_vig or "").strip(),
                "ESTATUS_PROCESO_HABILITACION": str(estatus or "").strip(),
            }

    raise HTTPException(status_code=404, detail="Conductor no encontrado por DNI en la columna E")

# ===== endpoints =====

@app.get("/health")
def health():
    return {"ok": True, "ts": int(time.time())}

@app.get("/driver")
def get_driver(
    doc: str = Query(..., description="DNI/CE exacto tal como aparece en la columna E"),
    t: str   = Query(..., description="token HMAC"),
    sheet_name: Optional[str] = Query(None, description="Nombre de hoja. Vacío = primera"),
    header_row: int = Query(12, description="Fila de encabezados, 1-based"),
):
    SECRET_KEY = os.getenv("SECRET_KEY", "").strip()
    ONEDRIVE_URL = os.getenv("ONEDRIVE_URL", "").strip()

    if not SECRET_KEY or not ONEDRIVE_URL:
        raise HTTPException(status_code=500, detail="Faltan variables de entorno: SECRET_KEY y/o ONEDRIVE_URL")

    # 1) verificar token
    if not verify(doc, t, SECRET_KEY):
        raise HTTPException(status_code=401, detail="token inválido")

    # 2) descargar Excel fresco
    url = od_to_download(ONEDRIVE_URL)
    url = f"{url}{'&' if '?' in url else '?'}_cb={int(time.time())}"
    r = requests.get(url, timeout=60, headers={"Cache-Control": "no-cache"})
    if r.status_code != 200:
        raise HTTPException(status_code=502, detail=f"No pude descargar Excel ({r.status_code})")

    # 3) parsear Excel sin pandas
    data = read_driver_from_excel(r.content, sheet_name, header_row, doc)
    return JSONResponse({"ok": True, "driver": data})

# NOTA: El Procfile en Render arrancará uvicorn/gunicorn como siempre.
